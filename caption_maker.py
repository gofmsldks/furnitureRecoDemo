from flask import Flask, request, render_template, redirect, url_for
import os
import pandas as pd
from openpyxl import load_workbook
from PIL import Image
import io
import base64
import openai
from langchain import LangChain, OpenAIClient
from dotenv import load_dotenv
from langchain.chains import RetrievalQA
from langchain_openai import OpenAI
from langchain_openai import ChatOpenAI
from langchain.prompts import PromptTemplate
from langchain.indexes import VectorstoreIndexCreator
from langchain_community.vectorstores import FAISS
from langchain.text_splitter import CharacterTextSplitter
from langchain.docstore.document import Document
from langchain.schema import HumanMessage
from langchain_openai import OpenAIEmbeddings
from dotenv import load_dotenv
from langchain_community.document_loaders import TextLoader
from langchain_text_splitters import CharacterTextSplitter
from langchain_core.runnables import RunnablePassthrough
from langchain_core.output_parsers import StrOutputParser
from langchain.chains import LLMChain


app = Flask(__name__)
load_dotenv()
# 결과 폴더 설정
RESULTS_FOLDER = 'results/'
app.config['RESULTS_FOLDER'] = RESULTS_FOLDER

# LangChain 및 OpenAI 클라이언트 설정


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/upload', methods=['POST'])
def upload_file():
    if request.method == 'POST':
        excel_file = request.files['excel_file']
        if excel_file:
            wb = load_workbook(excel_file)
            ws = wb.active
            # 엑셀에서 필요한 카테고리와 이미지 정보를 함께 추출
            data = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                data.append({
                    "seq": row[0],
                    "category": row[1],
                    "image": io.BytesIO(row[2])
                })

            results = []
            for item in data:
                img = Image.open(item["image"])
                temp_img_path = f'temp_image_{item["seq"]}.png'
                img.save(temp_img_path)

                # 캡션 생성
                caption = generate_caption(temp_img_path, item["category"])

                results.append({
                    'seq': item["seq"],
                    'image': temp_img_path,
                    'category': item["category"],
                    'caption': caption
                })

            # 결과를 엑셀 파일로 저장
            result_df = pd.DataFrame(results)
            result_df.to_excel(os.path.join(app.config['RESULTS_FOLDER'], 'output.xlsx'), index=False)

            return redirect(url_for('index'))


def generate_caption(image_path, category):
    # 이미지를 base64로 인코딩
    with open(image_path, "rb") as image_file:
        base64_image = base64.b64encode(image_file.read()).decode('utf-8')

    # GPT-4 Omni 모델로 요청을 보내기 위한 프롬프트 설정
    prompt = {
        "caption": f"Generate a detailed description and hashtags for a furniture image in the category {category}. Include aspects such as style, mood, and color scheme.",
        "image": base64_image
    }

    # GPT-4 Omni 모델 사용
    response = ChatOpenAI(
        model="gpt-4o",
        prompt=prompt,
        temperature=0.5,
        max_tokens=150
    )

    # 생성된 텍스트 추출
    return response.choices[0].text.strip()


if __name__ == '__main__':
    app.run(debug=True)
